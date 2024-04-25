from openpyxl import Workbook

# start by importing workbook

# example dictionary temperatures.
week_temperature = {
    'Monday': 54,
    'Tuesday': 68,
    'Wednesday': 62,
    'Thursday': 57,
    'Friday': 71,
}

workbook = Workbook()

# always close your worksheet before you run your program
worksheet = workbook.active

worksheet.title = 'Daily Temperatures'

# To print day and temp in the first rows
worksheet.cell(1, 1, 'Day')
worksheet.cell(1, 1, 'Temperature (F)')

row_index = 2

# here is a for loop to append the day and temp into
# different cells on the worksheet
for day, temp in week_temperature.items():
    worksheet.cell(row_index, 1, day)
    worksheet.cell(row_index, 2, temp)
    row_index += 1

workbook.save('temperatures.xlsx')