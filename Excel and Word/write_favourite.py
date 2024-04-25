from openpyxl import Workbook

favourite_foods = ['Pizza', 'Chocolate Cake', 'Brocolli']
favourite_colors = ['Blue', 'Purple', 'Green', 'Orange']

# you have to always do this to save the workbook
workbook = Workbook()

#
worksheet = workbook.active

# here is the worksheet title
worksheet.title = 'Favourite Things'

# this will append the text into the set cell.
worksheet.cell(1, 1, 'Favourite Foods')

# to print food in columns
for index, food in enumerate(favourite_foods):
    worksheet.cell(index + 1, 1, food)

# to print favourite color column
worksheet.cell(1, 2, 'Favourite colors')
for index, color in enumerate(favourite_colors):
    worksheet.cell(index + 2, 2, color)

# this is to save the file.
workbook.save('favourites.xlsx')






